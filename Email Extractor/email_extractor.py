"""
email_extractor.py — Inbox scraper for Interac e-Transfers + invoices + orders.

Logs into an IMAP mailbox (read-only, SSL-only), pulls messages within
a date range that look like Interac notifications, invoices/receipts,
or online order confirmations, parses each with the appropriate parser,
applies optional filters, and writes a single combined CSV + XLSX to
`output/`.

A `TYPE` column distinguishes `transfer` rows (Interac) from `invoice`
and `order` rows. Each row leaves the irrelevant columns blank for the
other types.

Security model (non-negotiable):
  * Password collected with getpass.getpass(); never echoed, never
    written to disk, never logged.
  * IMAP4_SSL on port 993 only. Plain IMAP4 is never imported.
  * INBOX selected with readonly=True; messages fetched with BODY.PEEK[]
    so the \\Seen flag is not set.
  * Server-side SEARCH narrows by date + sender substring + invoice
    subject keywords before any body fetch — no bulk inbox download.
  * No outbound HTTP. Hosted invoice URLs are recorded but never fetched.
  * Attached PDFs are parsed in memory and discarded — never written to
    disk. CSV + XLSX in output/ are the only artifacts written.

Run:
    python email_extractor.py
"""

import csv
import datetime
import email
import getpass
import imaplib
import os
import re
import sys
from decimal import Decimal
from email.message import Message
from typing import Callable, Iterable, Iterator, Optional, Union

from interac_parser import (
    ALLOWED_SENDERS,
    InteracRecord,
    parse_amount_filter,
    parse_interac_email,
)
from invoice_parser import (
    INLINE_INVOICE_SENDERS,
    SUBJECT_KEYWORDS,
    InvoiceRecord,
    parse_invoice_email,
)
from order_parser import (
    ORDER_SUBJECT_KEYWORDS,
    OrderRecord,
    parse_order_email,
)

OUTPUT_DIR = "output"

# Combined output schema. Transfer rows leave the invoice columns blank
# and vice versa. `TYPE` is "transfer", "invoice", or "order".
# Order rows reuse the invoice columns (INVOICE_NO carries the order #).
CSV_COLUMNS = [
    "TYPE", "DATE",
    # transfer-only
    "DIRECTION", "AMOUNT", "BANK", "CONTACT", "ACCOUNT", "REFERENCE", "MEMO",
    # invoice-only
    "VENDOR", "INVOICE_NO", "SUBTOTAL", "GST_HST", "QST_PST", "TOTAL", "CURRENCY",
    # shared
    "SOURCE", "LINK", "SUBJECT",
]

# Columns that carry monetary values in the XLSX; formatted as #,##0.00.
_MONEY_COLUMNS = ("AMOUNT", "SUBTOTAL", "GST_HST", "QST_PST", "TOTAL")

Record = Union[InteracRecord, InvoiceRecord, OrderRecord]


def _blank_row(for_csv: bool) -> dict:
    """Return a dict keyed by CSV_COLUMNS with blanks. CSV uses ""; XLSX
    uses None so number-formatted cells stay empty rather than showing 0."""
    placeholder = "" if for_csv else None
    return {c: placeholder for c in CSV_COLUMNS}


def _record_to_csv_row(rec: Record) -> dict:
    row = _blank_row(for_csv=True)
    if isinstance(rec, InteracRecord):
        row["TYPE"] = "transfer"
        row["SOURCE"] = "interac_email"
        row["LINK"] = ""
        row.update(rec.to_csv_row())
    else:
        row["TYPE"] = "order" if isinstance(rec, OrderRecord) else "invoice"
        row.update(rec.to_csv_row())
        row["AMOUNT"] = row["TOTAL"]  # mirror total into AMOUNT for unified sorting
    return row


def _record_to_xlsx_row(rec: Record) -> dict:
    row = _blank_row(for_csv=False)
    if isinstance(rec, InteracRecord):
        row["TYPE"] = "transfer"
        row["SOURCE"] = "interac_email"
        row["LINK"] = ""
        row.update(rec.to_excel_row())
    else:
        row["TYPE"] = "order" if isinstance(rec, OrderRecord) else "invoice"
        row.update(rec.to_excel_row())
        row["AMOUNT"] = row["TOTAL"]
    return row


# ---
# Inputs
# ---

class Inputs:
    __slots__ = (
        "email_address", "imap_host", "password",
        "since", "until",
        "contact_filter", "amount_predicate",
        "sender_allowlist",
        "subject_keywords",
    )

    def __init__(self, **kw):
        for k in self.__slots__:
            setattr(self, k, kw[k])


_HOST_HINTS = {
    "gmail.com":      "imap.gmail.com",
    "googlemail.com": "imap.gmail.com",
    "outlook.com":    "outlook.office365.com",
    "hotmail.com":    "outlook.office365.com",
    "live.com":       "outlook.office365.com",
    "yahoo.com":      "imap.mail.yahoo.com",
    "icloud.com":     "imap.mail.me.com",
    "me.com":         "imap.mail.me.com",
}


def _suggest_host(email_address: str) -> str:
    domain = email_address.partition("@")[2].lower().strip()
    return _HOST_HINTS.get(domain, "")


def _prompt(label: str, default: str = "") -> str:
    suffix = f" [{default}]" if default else ""
    raw = input(f"{label}{suffix}: ").strip()
    return raw or default


def _prompt_date(label: str, default: Optional[datetime.date] = None) -> datetime.date:
    default_str = default.isoformat() if default else ""
    while True:
        raw = _prompt(label, default_str)
        try:
            return datetime.date.fromisoformat(raw)
        except ValueError:
            print(f"  Invalid date {raw!r} — expected YYYY-MM-DD.")


def prompt_inputs() -> Inputs:
    print("Email extractor — Interac transfers + invoices + online orders")
    print("─────────────────────────────────────────────────────────────────")
    email_address = _prompt("Email address")
    if not email_address or "@" not in email_address:
        raise ValueError("Email address is required.")

    suggested_host = _suggest_host(email_address)
    imap_host = _prompt("IMAP host", suggested_host)
    if not imap_host:
        raise ValueError("IMAP host is required.")

    password = getpass.getpass("App password (input hidden): ")
    if not password:
        raise ValueError("Password is required.")

    print()
    since = _prompt_date("Since (YYYY-MM-DD)")
    until = _prompt_date("Until (YYYY-MM-DD)", default=datetime.date.today())
    if until < since:
        raise ValueError("'Until' cannot be earlier than 'Since'.")

    print()
    contact_filter = _prompt("Contact filter (optional substring, blank for all)")
    amount_filter_raw = _prompt(
        "Amount filter (optional, e.g. >=500, <=20, =1000, blank for all)"
    )
    try:
        amount_predicate = parse_amount_filter(amount_filter_raw)
    except ValueError as exc:
        raise ValueError(str(exc)) from None

    extras_raw = _prompt(
        "Additional sender domains (optional, comma-separated, e.g. rbc.com,scotiabank.com)"
    )
    extras = [e.strip() for e in extras_raw.split(",") if e.strip()]
    sender_allowlist = (
        list(ALLOWED_SENDERS)
        + list(INLINE_INVOICE_SENDERS)
        + [e if "@" in e or "*" in e else f"*@{e}" for e in extras]
    )

    return Inputs(
        email_address=email_address,
        imap_host=imap_host,
        password=password,
        since=since,
        until=until,
        contact_filter=contact_filter.lower() if contact_filter else "",
        amount_predicate=amount_predicate,
        sender_allowlist=sender_allowlist,
        subject_keywords=list(SUBJECT_KEYWORDS) + list(ORDER_SUBJECT_KEYWORDS),
    )


# ---
# IMAP layer
# ---

def connect_imap(host: str, user: str, password: str) -> imaplib.IMAP4_SSL:
    """Open an SSL IMAP connection on port 993 and select INBOX read-only."""
    try:
        M = imaplib.IMAP4_SSL(host, 993)
    except OSError as exc:
        raise SystemExit(f"Could not connect to {host}:993 — {exc}")

    try:
        M.login(user, password)
    except imaplib.IMAP4.error:
        # Never echo the password in error output.
        try:
            M.logout()
        except Exception:
            pass
        raise SystemExit(f"Auth failed for {user} on {host}.")

    typ, _ = M.select("INBOX", readonly=True)
    if typ != "OK":
        M.logout()
        raise SystemExit(f"Could not open INBOX (status: {typ}).")
    return M


def _imap_date(d: datetime.date) -> str:
    return d.strftime("%d-%b-%Y")


def _from_substrings(allowlist: list[str]) -> list[str]:
    """
    Reduce glob-style allowlist patterns to IMAP-friendly substrings.
    IMAP FROM is a substring search, so '*@bmo.com' becomes 'bmo.com'.
    """
    out: list[str] = []
    for pat in allowlist:
        s = pat.lower()
        if "@" in s:
            local, _, domain = s.partition("@")
            if "*" in local:
                out.append(domain.replace("*", ""))
            else:
                out.append(s)  # exact address
        else:
            out.append(s.replace("*", ""))
    # de-dupe while preserving order
    seen: set[str] = set()
    deduped: list[str] = []
    for s in out:
        s = s.strip(". ")
        if s and s not in seen:
            seen.add(s)
            deduped.append(s)
    return deduped


def _build_or_chain(criteria_pairs: list[tuple[str, str]]) -> str:
    """
    Build an IMAP OR chain. `criteria_pairs` is a list of (key, value) tuples.
    Returns a string like: 'OR FROM "a" OR FROM "b" FROM "c"'.
    Empty list returns ''.
    """
    if not criteria_pairs:
        return ""
    if len(criteria_pairs) == 1:
        k, v = criteria_pairs[0]
        return f'{k} "{v}"'
    # IMAP OR is binary; chain right-to-left.
    *rest, last = criteria_pairs
    chain = f'{last[0]} "{last[1]}"'
    for k, v in reversed(rest):
        chain = f'OR {k} "{v}" {chain}'
    return chain


def search_messages(M: imaplib.IMAP4_SSL,
                    since: datetime.date,
                    until: datetime.date,
                    sender_allowlist: list[str],
                    subject_keywords: list[str] | None = None) -> list[bytes]:
    """Server-side SEARCH narrowed by date range + (sender OR subject keyword)."""
    senders = _from_substrings(sender_allowlist)
    pairs: list[tuple[str, str]] = [("FROM", s) for s in senders]
    for kw in subject_keywords or []:
        pairs.append(("SUBJECT", kw))
    or_chain = _build_or_chain(pairs)

    # IMAP BEFORE is exclusive on internal date — bump by one day to make until inclusive.
    parts = [
        f"SINCE {_imap_date(since)}",
        f"BEFORE {_imap_date(until + datetime.timedelta(days=1))}",
    ]
    if or_chain:
        parts.append(f"({or_chain})")
    criteria = " ".join(parts)

    typ, data = M.search(None, criteria)
    if typ != "OK":
        raise SystemExit(f"IMAP SEARCH failed: {typ}")
    if not data or not data[0]:
        return []
    return data[0].split()


def fetch_and_parse(M: imaplib.IMAP4_SSL,
                    uids: Iterable[bytes]) -> Iterator[Record]:
    for uid in uids:
        typ, data = M.fetch(uid, "(BODY.PEEK[])")
        if typ != "OK" or not data or not isinstance(data[0], tuple):
            continue
        raw_bytes = data[0][1]
        msg = email.message_from_bytes(raw_bytes)

        # Dispatch order: Interac (narrowest scope) → Order (subject-scoped) →
        # Invoice (broadest fallback). Order is tried before Invoice so that
        # "Your order #..." style messages land as TYPE="order" instead of
        # being swept up by the invoice parser's broader subject scope.
        record: Optional[Record] = parse_interac_email(msg)
        if record is None:
            record = parse_order_email(msg)
        if record is None:
            record = parse_invoice_email(msg)
        if record is not None:
            yield record


# ---
# Filters & writer
# ---

def apply_filters(records: Iterable[Record],
                  contact_substring: str,
                  amount_predicate: Callable[[Decimal], bool]) -> Iterator[Record]:
    """Amount filter applies to all record types. Contact filter applies
    only to InteracRecord — invoices and orders are passed through (the
    user has explicitly asked for them and `vendor` is a different field)."""
    contact_substring = (contact_substring or "").lower()
    for r in records:
        if isinstance(r, InteracRecord):
            if contact_substring and contact_substring not in r.contact.lower():
                continue
            if not amount_predicate(r.amount):
                continue
        else:
            if not amount_predicate(r.total):
                continue
        yield r


_INVALID_FS_CHARS_RE = re.compile(r"[^A-Za-z0-9._\-]")


def _safe_for_path(s: str) -> str:
    return _INVALID_FS_CHARS_RE.sub("_", s)


def _output_paths(since: datetime.date, until: datetime.date) -> tuple[str, str]:
    """Return (csv_path, xlsx_path) for the given date range."""
    stem = _safe_for_path(f"email_records_{since.isoformat()}_{until.isoformat()}")
    return (
        os.path.join(OUTPUT_DIR, stem + ".csv"),
        os.path.join(OUTPUT_DIR, stem + ".xlsx"),
    )


def write_csv(records: Iterable[Record], path: str) -> int:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    count = 0
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for r in records:
            writer.writerow(_record_to_csv_row(r))
            count += 1
    return count


def write_xlsx(records: Iterable[Record], path: str) -> int:
    """Same rows as the CSV, but with native date / numeric typing."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise SystemExit(
            "openpyxl is required for XLSX output. Install with: pip install openpyxl"
        ) from exc

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Email Records"
    ws.append(CSV_COLUMNS)

    count = 0
    for r in records:
        row = _record_to_xlsx_row(r)
        ws.append([row[c] for c in CSV_COLUMNS])
        count += 1

    date_col = CSV_COLUMNS.index("DATE") + 1
    money_cols = [CSV_COLUMNS.index(c) + 1 for c in _MONEY_COLUMNS]
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=date_col).number_format = "yyyy-mm-dd"
        for col_idx in money_cols:
            ws.cell(row=row_idx, column=col_idx).number_format = "#,##0.00"

    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    ws.freeze_panes = "A2"
    for col in ws.columns:
        letter = col[0].column_letter
        width = max((len(str(c.value or "")) for c in col), default=10) + 2
        ws.column_dimensions[letter].width = min(max(width, 10), 50)

    wb.save(path)
    return count


def _confirm_overwrite(paths: list[str]) -> bool:
    existing = [p for p in paths if os.path.exists(p)]
    if not existing:
        return True
    listed = ", ".join(existing)
    answer = input(f"{listed} already exist(s). Overwrite? [y/N]: ").strip().lower()
    return answer in ("y", "yes")


# ---
# Main
# ---

def main() -> int:
    try:
        inputs = prompt_inputs()
    except (KeyboardInterrupt, EOFError):
        print("\nAborted.")
        return 130
    except ValueError as exc:
        print(f"Input error: {exc}")
        return 2

    csv_path, xlsx_path = _output_paths(inputs.since, inputs.until)
    if not _confirm_overwrite([csv_path, xlsx_path]):
        print("Aborted — existing file(s) preserved.")
        return 0

    print(f"\nConnecting to {inputs.imap_host} as {inputs.email_address} …")
    M = connect_imap(inputs.imap_host, inputs.email_address, inputs.password)

    try:
        print(f"Searching {inputs.since.isoformat()} → {inputs.until.isoformat()} …")
        uids = search_messages(
            M, inputs.since, inputs.until,
            inputs.sender_allowlist, inputs.subject_keywords,
        )
        print(f"  {len(uids)} candidate messages on the server.")

        parsed = list(fetch_and_parse(M, uids))
        n_transfers = sum(1 for r in parsed if isinstance(r, InteracRecord))
        n_orders    = sum(1 for r in parsed if isinstance(r, OrderRecord))
        n_invoices  = sum(1 for r in parsed if isinstance(r, InvoiceRecord))
        print(f"  {n_transfers} parsed as Interac transfers, "
              f"{n_orders} as orders, {n_invoices} as invoices.")

        filtered = list(apply_filters(parsed, inputs.contact_filter, inputs.amount_predicate))
        print(f"  {len(filtered)} after filters.")

        csv_count = write_csv(filtered, csv_path)
        xlsx_count = write_xlsx(filtered, xlsx_path)
        print(f"\n{csv_count} records written → {csv_path}")
        print(f"{xlsx_count} records written → {xlsx_path}")
        return 0
    finally:
        try:
            M.close()
        except Exception:
            pass
        try:
            M.logout()
        except Exception:
            pass


if __name__ == "__main__":
    sys.exit(main())

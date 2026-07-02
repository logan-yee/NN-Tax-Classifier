"""Two-stage inference: data/incoming.csv -> output/classified_<date>.xlsx.

Output is a COPY of the master workbook with the pivot/summary sheets dropped and
a single "Predictions" sheet appended (the only addition to the master format).
Vendor rows are highlighted amber; unresolved e-transfers and low-confidence rows
are highlighted in a distinct red.
"""
from __future__ import annotations
import shutil
from datetime import date
from pathlib import Path

import numpy as np
import pandas as pd
import torch
from openpyxl.styles import Font, PatternFill
from torch.utils.data import DataLoader

from .config import (
    INCOMING_CSV, OUTPUT_DIR, STAGE1_DIR, STAGE2_DIR, MASTER_XLSX,
    BATCH_SIZE, CONFIDENCE_THRESHOLD, HIGHLIGHT_FILL_COLOR, UNRESOLVED_FILL_COLOR,
    KNOWN_CONTACTS_CSV, DROP_SHEETS, PREDICTIONS_SHEET, resolve_interac_sources,
)
from .dataset import TxnInferenceDataset
from .model import load_model
from .rules import (
    apply_rules, vendor_highlight_mask, load_etransfer_index, load_known_contacts,
)
from .schema import MASTER_COLUMNS, conform_to_master


def _infer(model, loader, device) -> tuple[np.ndarray, np.ndarray]:
    model.to(device).eval()
    all_preds, all_confs = [], []
    with torch.no_grad():
        for batch in loader:
            batch = {k: v.to(device) for k, v in batch.items()}
            logits = model(**batch)["logits"]
            probs = torch.softmax(logits, dim=-1)
            conf, pred = probs.max(dim=-1)
            all_preds.append(pred.cpu().numpy())
            all_confs.append(conf.cpu().numpy())
    return np.concatenate(all_preds), np.concatenate(all_confs)


def _write_xlsx(df: pd.DataFrame, out_path: Path,
                vendor_mask: np.ndarray, unresolved_mask: np.ndarray) -> None:
    """Write the master-copy workbook: drop pivot/summary sheets, append a single
    Predictions sheet. Unresolved rows (red) take precedence over vendor rows (amber)."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    vendor_fill = PatternFill("solid", fgColor=HIGHLIGHT_FILL_COLOR)
    unresolved_fill = PatternFill("solid", fgColor=UNRESOLVED_FILL_COLOR)

    # Start from a copy of the master so its data sheets are preserved verbatim.
    if MASTER_XLSX.exists():
        shutil.copyfile(MASTER_XLSX, out_path)
        writer_kwargs = dict(engine="openpyxl", mode="a", if_sheet_exists="replace")
    else:  # fallback: standalone Predictions-only workbook
        writer_kwargs = dict(engine="openpyxl")

    with pd.ExcelWriter(out_path, **writer_kwargs) as w:
        if writer_kwargs.get("mode") == "a":
            for s in DROP_SHEETS:
                if s in w.book.sheetnames:
                    del w.book[s]
        df.to_excel(w, sheet_name=PREDICTIONS_SHEET, index=False)
        ws = w.sheets[PREDICTIONS_SHEET]
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")
        for col in ws.columns:
            width = min(50, max((len(str(c.value)) for c in col if c.value is not None), default=10) + 2)
            ws.column_dimensions[col[0].column_letter].width = width
        # Highlight: unresolved (red) wins over vendor (amber). Applied by position;
        # row order is preserved from df so masks stay aligned.
        ncols = len(df.columns)
        for offset in range(len(df)):
            fill = unresolved_fill if unresolved_mask[offset] else (
                vendor_fill if vendor_mask[offset] else None)
            if fill is not None:
                excel_row = offset + 2  # +1 header, +1 to 1-based
                for c in range(1, ncols + 1):
                    ws.cell(row=excel_row, column=c).fill = fill


def predict() -> Path:
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    df = pd.read_csv(INCOMING_CSV)
    df = conform_to_master(df)
    print(f"[predict] loaded {len(df)} rows from {INCOMING_CSV}")

    stage1, tok1, meta1 = load_model(STAGE1_DIR)
    ds1 = TxnInferenceDataset(df, tok1)
    loader1 = DataLoader(ds1, batch_size=BATCH_SIZE)
    preds1, confs1 = _infer(stage1, loader1, device)
    id_to_label1 = {int(k): v for k, v in meta1["id_to_label"].items()}
    df["Category"] = [id_to_label1[int(p)] for p in preds1]
    df["Confidence"] = confs1.astype(float)
    print(f"[predict] stage1 done; class counts: {df['Category'].value_counts().to_dict()}")

    df["Sub-Category1"] = None
    df["Confidence_Stage2"] = np.nan
    expense_mask = df["Category"] == "EXPENSES"
    if expense_mask.any():
        stage2, tok2, meta2 = load_model(STAGE2_DIR)
        ds2 = TxnInferenceDataset(df[expense_mask].reset_index(drop=True), tok2)
        loader2 = DataLoader(ds2, batch_size=BATCH_SIZE)
        preds2, confs2 = _infer(stage2, loader2, device)
        id_to_label2 = {int(k): v for k, v in meta2["id_to_label"].items()}
        df.loc[expense_mask, "Sub-Category1"] = [id_to_label2[int(p)] for p in preds2]
        df.loc[expense_mask, "Confidence_Stage2"] = confs2.astype(float)
        print(f"[predict] stage2 done on {expense_mask.sum()} EXPENSES rows")

    # Additive rule layer: vendor labels + e-transfer cross-reference + category chain.
    # Account_Lean and Unresolved are computed internally and left out of out_cols.
    interac_sources = resolve_interac_sources()
    et_index = load_etransfer_index(interac_sources)
    known = load_known_contacts(KNOWN_CONTACTS_CSV)
    print(f"[predict] e-transfer xref: {len(et_index)} records from "
          f"{[p.name for p in interac_sources]}; {len(known)} known contacts")
    df = apply_rules(df, et_index=et_index, known_contacts=known)
    print(f"[predict] decisions: {df['Decision_Source'].value_counts().to_dict()}")

    # Highlight masks (computed before trimming internal columns; row order preserved).
    conf_mask = (df["Confidence"] < CONFIDENCE_THRESHOLD) | (
        (df["Category"] == "EXPENSES") & (df["Confidence_Stage2"] < CONFIDENCE_THRESHOLD)
    )
    unresolved_mask = (df["Unresolved"].fillna(False).to_numpy() | conf_mask.to_numpy())
    vendor_mask = vendor_highlight_mask(df).to_numpy()
    print(f"[predict] highlight: {int(vendor_mask.sum())} vendor / "
          f"{int(unresolved_mask.sum())} unresolved rows")

    out_cols = MASTER_COLUMNS + ["Confidence", "Confidence_Stage2", "Decision_Source"]
    df = df[out_cols]
    out_path = OUTPUT_DIR / f"classified_{date.today().isoformat()}.xlsx"
    _write_xlsx(df, out_path, vendor_mask, unresolved_mask)
    print(f"[predict] wrote -> {out_path}")
    return out_path


if __name__ == "__main__":
    predict()

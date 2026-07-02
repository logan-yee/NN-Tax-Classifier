"""Tests for the ADDITIVE predict-time rule layer (src/rules.py).

Run from the Neural Classifier/ directory:
    python -m pytest src/test_rules.py -v

These tests exercise the rule logic directly (no torch / no trained model), and
a small openpyxl round-trip to confirm the vendor highlight actually renders.
"""
from __future__ import annotations

import pandas as pd
import pytest

from src import config
from src.rules import (
    apply_rules, vendor_highlight_mask, load_known_contacts, load_etransfer_index,
    match_etransfer, contact_lean,
)
from src.schema import MASTER_COLUMNS


def make_df(rows: list[dict]) -> pd.DataFrame:
    """Build a MASTER-shaped frame from partial row dicts (missing cols -> NA)."""
    df = pd.DataFrame({c: [r.get(c, pd.NA) for r in rows] for c in MASTER_COLUMNS})
    return df


def one(rows: dict) -> pd.Series:
    return apply_rules(make_df([rows])).iloc[0]


# --------------------------------------------------------------------------- #
# Category chain + vendor label cases
# --------------------------------------------------------------------------- #

def test_td_generic():
    """TD generic -> Advertising / Sub-Cat1 unchanged / Sub-Cat2 unchanged / source_td / business."""
    r = one({
        "Source": "TD_Visa", "Description": "generic purchase downtown",
        "Category": "PERSONAL", "Sub-Category1": "shopping",  # model output
    })
    assert r["Category"] == "Advertising"
    assert r["Sub-Category1"] == "shopping"          # unchanged
    assert pd.isna(r["Sub-Category2"])               # unchanged (was NA)
    assert r["Decision_Source"] == "source_td"
    assert r["Account_Lean"] == "business"
    assert not vendor_highlight_mask(make_df([{"Sub-Category2": r["Sub-Category2"]}]))[0]


def test_td_amazon_td_wins_and_vendor_label():
    """TD Amazon -> Advertising (TD wins) / Sub-Category2=AMAZON / highlighted."""
    df = apply_rules(make_df([{
        "Source": "TD_Visa", "Description": "AMAZON.CA*RT4 order",
        "Category": "PERSONAL", "Sub-Category1": "shopping",
    }]))
    r = df.iloc[0]
    assert r["Category"] == "Advertising"            # TD outranks everything
    assert r["Sub-Category2"] == "AMAZON"
    assert r["Sub-Category1"] == "shopping"          # untouched
    assert r["Decision_Source"] == "source_td"
    assert bool(vendor_highlight_mask(df)[0]) is True


def test_bmo_car_charge():
    """BMO car charge -> Business Expense / keyword_car / business."""
    r = one({
        "Source": "BMO_MC", "Description": "car insurance monthly",
        "Category": "PERSONAL", "Sub-Category1": pd.NA,
    })
    assert r["Category"] == "Business Expense"
    assert r["Decision_Source"] == "keyword_car"
    assert r["Account_Lean"] == "business"


def test_tangerine_paypal():
    """Tangerine PayPal -> Sub-Category2=PAYPAL / Sub-Cat1 unchanged / personal lean / highlighted."""
    df = apply_rules(make_df([{
        "Source": "Tangerine", "Description": "PAYPAL *MERCHANT123",
        "Category": "PERSONAL", "Sub-Category1": "groceries",
    }]))
    r = df.iloc[0]
    assert r["Sub-Category2"] == "PAYPAL"
    assert r["Sub-Category1"] == "groceries"         # unchanged
    assert r["Account_Lean"] == "personal"
    assert r["Decision_Source"] == "source_prior"    # Tangerine prior
    assert bool(vendor_highlight_mask(df)[0]) is True


def test_bmo_facebook_charge():
    """BMO facebook charge -> Advertising / keyword_facebook."""
    r = one({
        "Source": "BMO_CK", "Description": "facebook ads campaign",
        "Category": "PERSONAL", "Sub-Category1": pd.NA,
    })
    assert r["Category"] == "Advertising"
    assert r["Decision_Source"] == "keyword_facebook"
    assert r["Account_Lean"] == "business"


def test_tangerine_grocery_unchanged_path():
    """Tangerine grocery -> personal / source_prior (Category kept = model output)."""
    r = one({
        "Source": "Tangerine", "Description": "grocery store loblaws",
        "Category": "PERSONAL", "Sub-Category1": "groceries",
    })
    assert r["Category"] == "PERSONAL"               # model Category kept
    assert r["Account_Lean"] == "personal"
    assert r["Decision_Source"] in ("source_prior", "model")
    assert pd.isna(r["Sub-Category2"])


def test_contact_map_only_when_no_higher_rule():
    """Contact row, no other match -> contact_map."""
    r = one({
        "Source": "CIBC", "Description": "shopify payout deposit",
        "Category": "REVENUE", "Sub-Category1": pd.NA,
    })
    assert r["Decision_Source"] == "contact_map"
    assert r["Account_Lean"] == "business"           # "shopify" -> expense
    assert r["Category"] == "REVENUE"                # Category kept


# --------------------------------------------------------------------------- #
# Regression: a row matching NO new rule is byte-for-byte unchanged
# --------------------------------------------------------------------------- #

def test_regression_no_rule_fires_preserves_everything():
    """No rule matches -> Category, Sub-Category1 AND Sub-Category2 all preserved, Decision=model."""
    base = {
        "Source": "CIBC", "Description": "interac e-transfer from john",
        "Category": "REVENUE", "Sub-Category1": "REVENUE", "Sub-Category2": "PRESET_VALUE",
    }
    r = one(base)
    assert r["Category"] == "REVENUE"                # unchanged
    assert r["Sub-Category1"] == "REVENUE"           # unchanged
    assert r["Sub-Category2"] == "PRESET_VALUE"      # unchanged (no vendor match)
    assert r["Decision_Source"] == "model"
    # not a vendor label -> not highlighted
    assert bool(vendor_highlight_mask(make_df([base]))[0]) is False


# --------------------------------------------------------------------------- #
# Highlight renders in a real xlsx (openpyxl round-trip)
# --------------------------------------------------------------------------- #

def test_vendor_highlight_renders_in_xlsx(tmp_path):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    df = apply_rules(make_df([
        {"Source": "Tangerine", "Description": "PAYPAL purchase", "Category": "PERSONAL"},
        {"Source": "CIBC", "Description": "plain deposit", "Category": "REVENUE"},
    ]))
    mask = vendor_highlight_mask(df).to_numpy()
    assert mask.tolist() == [True, False]

    out = tmp_path / "hl.xlsx"
    fill = PatternFill("solid", fgColor=config.HIGHLIGHT_FILL_COLOR)
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="Predictions", index=False)
        ws = w.sheets["Predictions"]
        ncols = len(df.columns)
        for offset, is_vendor in enumerate(mask):
            if is_vendor:
                for c in range(1, ncols + 1):
                    ws.cell(row=offset + 2, column=c).fill = fill

    wb = load_workbook(out)
    ws = wb["Predictions"]
    # row 2 = PayPal (highlighted), row 3 = plain (not)
    assert ws.cell(row=2, column=1).fill.fgColor.rgb.endswith(config.HIGHLIGHT_FILL_COLOR)
    plain = ws.cell(row=3, column=1).fill
    assert plain.patternType in (None, "none")


# --------------------------------------------------------------------------- #
# E-transfer cross-reference
# --------------------------------------------------------------------------- #

def et_index(rows):
    """Build a normalized e-transfer index frame (date, direction, amount, contact)."""
    return pd.DataFrame([
        {"date": pd.Timestamp(d), "direction": dirn, "amount": round(amt, 2), "contact": c}
        for d, dirn, amt, c in rows
    ])


def one_xref(row, index, known=None):
    return apply_rules(make_df([row]), et_index=index, known_contacts=known or {}).iloc[0]


def test_etransfer_business_sent_to_expenses_via_suffix():
    """Sent to a company (INC suffix) -> EXPENSES / etransfer_xref / Note carries the name."""
    idx = et_index([("2026-04-04", "sent", 888.88, "CONSULTATION SDV D2V SCIENCE INC")])
    r = one_xref({
        "Source": "BMO_CK", "Description": "INTERACe-TransferSent",
        "DATE": "2026-04-04", "Amount": 888.88, "Debit": 888.88, "Category": "PERSONAL",
    }, idx)
    assert r["Category"] == "EXPENSES"
    assert r["Account_Lean"] == "business"
    assert r["Decision_Source"] == "etransfer_xref"
    assert bool(r["Unresolved"]) is False
    assert "CONSULTATION SDV D2V SCIENCE INC" in str(r["Note"])


def test_etransfer_business_received_to_revenue():
    """Received from a business contact -> REVENUE."""
    idx = et_index([("2026-04-06", "received", 1500.0, "Acme Holdings Ltd")])
    r = one_xref({
        "Source": "BMO_CK", "Description": "INTERACe-TransferReceived",
        "DATE": "2026-04-06", "Amount": 1500.0, "Credit": 1500.0, "Category": "PERSONAL",
    }, idx)
    assert r["Category"] == "REVENUE"
    assert r["Decision_Source"] == "etransfer_xref"


def test_etransfer_personal_known_contact():
    """Known personal contact -> PERSONAL regardless of direction."""
    idx = et_index([("2026-04-08", "sent", 50.0, "Mom")])
    r = one_xref({
        "Source": "Tangerine", "Description": "INTERACe-TransferSent",
        "DATE": "2026-04-08", "Amount": 50.0, "Debit": 50.0, "Category": "EXPENSES",
    }, idx, known={"mom": "personal"})
    assert r["Category"] == "PERSONAL"
    assert r["Account_Lean"] == "personal"
    assert r["Decision_Source"] == "etransfer_xref"


def test_etransfer_contact_found_but_unclassifiable_is_unresolved():
    """Contact recovered but no map/suffix hit -> unresolved, model kept, name in Note."""
    idx = et_index([("2026-04-10", "sent", 350.0, "Jayden Lam")])
    r = one_xref({
        "Source": "BMO_CK", "Description": "INTERACe-TransferSent",
        "DATE": "2026-04-10", "Amount": 350.0, "Debit": 350.0, "Category": "PERSONAL",
    }, idx)
    assert r["Decision_Source"] == "etransfer_unresolved"
    assert bool(r["Unresolved"]) is True
    assert r["Category"] == "PERSONAL"               # model output kept
    assert "Jayden Lam" in str(r["Note"])


def test_etransfer_no_match_is_unresolved():
    """Transfer with no email match -> unresolved, no Note tag."""
    idx = et_index([("2026-04-10", "sent", 999.99, "Someone Else")])
    r = one_xref({
        "Source": "BMO_CK", "Description": "INTERACe-TransferSent",
        "DATE": "2026-04-10", "Amount": 222.22, "Debit": 222.22, "Category": "PERSONAL",
    }, idx)
    assert r["Decision_Source"] == "etransfer_unresolved"
    assert bool(r["Unresolved"]) is True
    assert pd.isna(r["Note"]) or "xref" not in str(r["Note"])


def test_etransfer_outside_date_window_no_match():
    """Same amount+direction but >window days away must NOT match."""
    idx = et_index([("2026-01-01", "sent", 120.0, "CONSULTATION SDV D2V SCIENCE INC")])
    r = one_xref({
        "Source": "BMO_CK", "Description": "INTERACe-TransferSent",
        "DATE": "2026-04-10", "Amount": 120.0, "Debit": 120.0, "Category": "PERSONAL",
    }, idx)
    assert r["Decision_Source"] == "etransfer_unresolved"
    assert bool(r["Unresolved"]) is True


def test_match_etransfer_and_contact_lean_units():
    idx = et_index([("2026-04-04", "sent", 888.88, "Foo Bar Inc")])
    assert match_etransfer(idx, pd.Timestamp("2026-04-05"), 888.88, "sent") == "Foo Bar Inc"
    assert match_etransfer(idx, pd.Timestamp("2026-04-05"), 888.88, "received") is None
    assert contact_lean("Foo Bar Inc", {}) == "business"          # suffix
    assert contact_lean("Daniel Mosqueda", {"daniel mosqueda": "business"}) == "business"
    assert contact_lean("Random Person", {}) is None              # unresolved


def test_load_known_contacts_skips_comments():
    contacts = load_known_contacts(config.KNOWN_CONTACTS_CSV)
    assert "sdv d2v science" in contacts
    assert contacts["sdv d2v science"] == "business"


def test_load_etransfer_index_reads_smoke_csv():
    sources = config.resolve_interac_sources()
    assert sources, "no interac source resolved"
    idx = load_etransfer_index(sources)
    assert len(idx) > 0
    assert set(["date", "direction", "amount", "contact"]).issubset(idx.columns)
    assert set(idx["direction"].unique()).issubset({"sent", "received", "not a transaction"})


# --------------------------------------------------------------------------- #
# Output workbook: master data sheets preserved, single Predictions sheet, 2 colours
# --------------------------------------------------------------------------- #

def test_output_workbook_structure_and_two_colour_highlight(tmp_path):
    import numpy as np
    from openpyxl import load_workbook
    from src.predict import _write_xlsx

    if not config.MASTER_XLSX.exists():
        pytest.skip("master workbook not available")

    out_cols = MASTER_COLUMNS + ["Confidence", "Confidence_Stage2", "Decision_Source"]
    df = pd.DataFrame({c: [pd.NA, pd.NA, pd.NA] for c in out_cols})
    df["Source"] = ["Tangerine", "BMO_CK", "CIBC"]
    df["Sub-Category2"] = ["PAYPAL", pd.NA, pd.NA]          # row0 vendor
    vendor_mask = np.array([True, False, False])
    unresolved_mask = np.array([False, True, False])        # row1 unresolved (red wins)

    out = tmp_path / "classified_test.xlsx"
    _write_xlsx(df, out, vendor_mask, unresolved_mask)

    wb = load_workbook(out)
    # pivots dropped, data sheets preserved, single Predictions sheet
    for s in config.DROP_SHEETS:
        assert s not in wb.sheetnames
    assert "MASTER 2025" in wb.sheetnames
    assert "ETRANSFER" in wb.sheetnames
    assert config.PREDICTIONS_SHEET in wb.sheetnames
    assert not any("Review" in s for s in wb.sheetnames)

    ws = wb[config.PREDICTIONS_SHEET]
    assert ws.cell(row=2, column=1).fill.fgColor.rgb.endswith(config.HIGHLIGHT_FILL_COLOR)   # vendor amber
    assert ws.cell(row=3, column=1).fill.fgColor.rgb.endswith(config.UNRESOLVED_FILL_COLOR)  # unresolved red
    assert ws.cell(row=4, column=1).fill.patternType in (None, "none")                       # plain


if __name__ == "__main__":
    raise SystemExit(pytest.main([__file__, "-v"]))
